import glob
import shutil

import matplotlib.pyplot as plt
import openpyxl as op
import pywt
import numpy as np
import os

from sklearn.cluster import KMeans

workbook = op.load_workbook("C:\\Users\\ivmi0322\\Desktop\\OilCoreResearch.xlsx")
worksheet = workbook.active

# Get values from Excel file
x = []
platforms = {}
counter = 0
for row in worksheet.iter_rows(2, worksheet.max_row):
    x.append(row[1].value)
for col in worksheet.iter_cols(3, worksheet.max_column):
    counter += 1
    first_row_flag = True
    values = []
    name = col[0].value
    for row in range(1, worksheet.max_row):
        values.append(col[row].value)
    platforms.update({str(counter) + " - " + name: values})
print(len(platforms))

# Plot dependencies as 2D grafs
counter = 0
for key, value in platforms.items():
    counter += 1
    plt.plot(x, value)
    plt.grid()
    plt.title('Распределение пор по размерам - ' + key)
    plt.xscale('log')
    plt.ylabel('Объем пор , %')
    plt.xlabel('Радиус пор , мкм')
    plt.savefig("resources//grafs//" + key + ".png")
    plt.clf()

# Calculating wavelet scalograms
counter = 0
vmax = 0
for key, value in platforms.items():
    widths = np.arange(1, 30)
    cwtmatr, freqs = pywt.cwt(value, widths, 'mexh')
    vmaxTemp = abs(cwtmatr).max()
    if vmax < vmaxTemp:
        vmax = vmaxTemp
features = []
for key, value in platforms.items():
    plt.xscale('linear')
    counter += 1
    widths = np.arange(1, 30)
    cwtmatr, freqs = pywt.cwt(value, widths, 'mexh')
    plt.imshow(cwtmatr, extent=[0, max(x), 1, 31], cmap='PRGn', aspect='auto',
               vmax=vmax, vmin=-vmax)
    # plt.ylabel('frequency , Hz')
    # plt.xlabel('time , sec')
    # plt.title(' Вейвлет скалограмма - ' + key)
    # plt.colorbar()
    plt.savefig("resources//wavelets//" + key + ".png")
    plt.clf()

#     this logic is for clustering features from wavelets directly
#     cwtmatrFlatted = cwtmatr.flatten()
#     features.append(cwtmatrFlatted)
# features = np.array(features)
# features.reshape(-1, 1711)
# print(features.shape)
# # cluster feature vectors
# kmeans = KMeans(n_clusters=10, random_state=22)
# label = kmeans.fit_predict(features)
#
# # Getting unique labels
#
# u_labels = np.unique(label)
#
# filenames = np.array(list(platforms.keys()))
# # holds the cluster id and the images { id: [images] }
# groups = {}
# for file, cluster in zip(filenames, kmeans.labels_):
#     if cluster not in groups.keys():
#         groups[cluster] = []
#         groups[cluster].append(file)
#     else:
#         groups[cluster].append(file)
#
# files = glob.glob('C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\groups\\wavelets\\*')
# for f in files:
#     os.remove(f)
# files = glob.glob('C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\groups\\grafs\\*')
# for f in files:
#     os.remove(f)
# for group in groups.keys():
#     directory = str(group)
#
#     # Parent Directories
#     parent_wavelets_dir = "C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\groups\\wavelets\\"
#     parent_grafs_dir = "C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\groups\\grafs\\"
#
#     # Path
#     path_wavelets = os.path.join(parent_wavelets_dir, directory)
#     path_grafs = os.path.join(parent_grafs_dir, directory)
#
#     # Create the directory
#     os.makedirs(path_wavelets)
#     os.makedirs(path_grafs)
#     for image in groups[group]:
#         shutil.copy2('C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\wavelets\\' + image + '.png',
#                      path_wavelets)
#         shutil.copy2('C:\\Users\\ivmi0322\\PycharmProjects\\OilCoreResearch\\resources\\grafs\\' + image + '.png',
#                      path_grafs)
